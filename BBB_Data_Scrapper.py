from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook
from openpyxl import load_workbook
import time
import json
import os

def update_progress(progress_data):

    with open('progress.json', 'w') as f:

        f.write(json.dumps(progress_data))

def load_progress_tracker():

    progress_status, progress = 'New', {}

    if os.path.exists('progress.json'):

        user_input = input("Do you want to continue old scraping process?\nEnter (Yes/No): ")

        if user_input in ['Yes', 'No']:
            
            if 'Yes' == user_input:

                try:

                    with open('progress.json', 'r') as f:

                        progress = json.loads(f.read())

                    progress_status = "Old"
                    
                except Exception as e:

                    print("Error in reading progress.json")    
    
        else:

            print("Invalid Input.")
            
            progress_status = "Stop"
    
    return progress_status, progress

def get_user_input():

    query = input("Enter Your Query: ")

    location = input("Enter Location: ")

    country = input("Enter Country (USA/CA): ")

    if country not in ["USA", "CA"]:
        print("Invalid Country.")
        exit()

    return [query, location, country]

def get_total_pages(driver):

    total_pages = 1

    try:
        
        pagination = driver.find_element(By.XPATH, "//nav[@aria-label='pagination']")

        last_page_tag = pagination.find_elements(By.TAG_NAME, "a")[-3]


        total_pages = int(last_page_tag.text.split("\n")[-1])

    except Exception as e:

        pass

    return total_pages

def get_business_details(driver, link, link_attrb, remaining_businesses):

    print("")

    business_data_dict = {"Title": '',"Overview": '', "Product & Services": '',"Business Started": '',  "Hours of Operation" : '', "Contact Information" : '', "Business Categories" : '', 'Website': '', "Phone Number":''}

    is_all_extracted = False

    try:
        
        business_data_dict["Title"] = link.text

        print(f"Scraping business '{link.text}' - {remaining_businesses} more businesses until next page.")

        driver.switch_to.window(driver.window_handles[1]) #switch to second tab
        
        # Overview and Product & Services

        driver.get(link_attrb)

        time.sleep(10)

        try:

            overview = driver.find_element(By.XPATH, "//div[contains(@class,'dtm-overview')]//div").text

            if 'More Info on Local BBB' in overview:
                business_data_dict["Overview"] = ''    
            else:
                business_data_dict["Overview"] = overview

        except:

            print("Overview not found.")

        try:
            products_services = driver.find_element(By.XPATH, "//div[contains(@class,'dtm-products-services')]").text
            business_data_dict["Product & Services"] = products_services
        except:
            print("Products & Services not found.")


        driver.get(link_attrb + "/details")

        time.sleep(10)

        # Business Started, Hours of Operation, All contact person, Business categories

        try:
            div =  driver.find_element(By.XPATH, "//div[@class='stack css-n8vred e1ri33r70']")

            data_heading = div.find_elements(By. TAG_NAME, "dt" )

            data_value = div.find_elements(By. TAG_NAME, "dd")

            for index in range(len(data_heading)):

                try:
                    label = data_heading[index].text
                    if label in required_data_list:
                        if label.endswith(":"):
                            label = label[:-1]

                        value = data_value[index].text
                        
                        business_data_dict[label] = value
                except:

                    # print("Error in heading loop")
                    pass
        
        except Exception as e:

            print("Error in extracting business details.")

        # Contact Info
        ## Phone Number, Website

        try:
            contact_div = driver.find_element(By.XPATH, "//div[contains(@class,'dtm-contact')]")

            try:
                website = contact_div.find_element(By.TAG_NAME, "a")
                business_data_dict["Website"] = website.get_attribute('href')
            except:
                print("Website not found.")

            try:
                phone_number = contact_div.find_element(By.XPATH, "//a[contains(@class,'dtm-phone')]")
                business_data_dict["Phone Number"] = phone_number.text
            except:
                print("Phone Number not found.")

            is_all_extracted = True

        except:

            print("Contact info not found.")

        
    except:
        
        print("Skipping:", link_attrb)

    return is_all_extracted, business_data_dict

if __name__ == "__main__":

    chrome_options = Options()

    chrome_options.add_argument('--headless')

    chrome_options.add_argument("--log-level=3")

    chrome_options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36 Edg/124.0.0.0')

    driver = webdriver.Chrome(options=chrome_options)

    progress_status, progress = load_progress_tracker()

    wb = None

    ws = None

    query, location, country = '', '', ''

    required_data_list = ["Business Started:", "Hours of Operation", "Contact Information", "Business Categories"]

    total_pages = 0

    # Setup

    if progress_status == "Stop":

        exit()
    
    elif progress_status == "New":

        wb = Workbook() # Workbook

        ws = wb.active # Worksheet

        progress = {'query': '', 'location': '', 'country': '', 'page_no': 1, 'total_pages' : '', 'visited_businesses': []}

        headers = ["Title", "Overview", "Product & Services", "Business Started",  "Hours of Operation", "Contact Information", "Business Categories", 'Website', "Phone Number"]

        ws.append(headers) # Write headers in Worksheet

        wb.save('Output.xlsx') # Save Excel File

        query, location, country = get_user_input() # Get User Input

        progress['query'] = query

        progress['location'] = location
        
        progress['country'] = country

    elif progress_status == "Old":

        wb = load_workbook('Output.xlsx')

        ws = wb.active # Worksheet

        query = progress['query']

        location = progress['location']

        country = progress['country']

    # Scraping Process

    url = f"https://www.bbb.org/search?find_country={country}&find_loc={location}&find_text={query}&page="

    if progress_status == "New":
    
        driver.get(url+"1")

        time.sleep(10)

        total_pages = get_total_pages(driver)

        progress["total_pages"] = total_pages
    
    elif progress_status == "Old":

        total_pages = progress["total_pages"]

    driver.execute_script(f'''window.open("","_blank");''') # Open new tab

    print("")

    print("Total PAGES:", total_pages)

    for pageNumber in range(1, total_pages+1):
        
        if pageNumber < progress["page_no"]:

            continue

        print("")

        print("PAGE#", pageNumber, f"of pages {total_pages}")

        progress["page_no"] = pageNumber

        driver.switch_to.window(driver.window_handles[0]) #switch to first tab
        
        while True:

            driver.get(url + str(pageNumber))

            time.sleep(10)
            
            if "500 Error" not in driver.page_source:
                break

        business_links = driver.find_elements(By.XPATH, "//a[@class='text-blue-medium css-1jw2l11 eou9tt70']")

        remaining_businesses_of_page = len(business_links)

        for link in business_links:

            remaining_businesses_of_page -= 1

            driver.switch_to.window(driver.window_handles[0]) #switch to first tab
            
            link_attrb = link.get_attribute("href")
            
            if link_attrb.startswith("https://www.bbb.org/") and link_attrb not in progress['visited_businesses']:

                is_all_extracted, business = get_business_details(driver, link, link_attrb, remaining_businesses_of_page)

                if is_all_extracted:

                    progress['visited_businesses'].append(link_attrb)

                    ws.append(list(business.values()))

                    wb.save('Output.xlsx')

                    update_progress(progress)
        
        progress['visited_businesses'] = []
    

    # Remove progress file if scraping process done

    try:
        os.remove('progress.json')
    except Exception as e:
        print("Error in removing progress.json.")